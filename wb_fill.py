# wb_fill.py
import json
import random
import re
from pathlib import Path
from openpyxl import load_workbook

TITLE_MAX = 60
DESC_MAX = 2000


# ---------- Themes (QSS) ----------
DEFAULT_THEMES = {
    "Midnight": """
        QWidget { background:#0b1220; color:#e7eefc; font-size:13px; }
        QLabel#Title { font-size:22px; font-weight:800; }
        QLabel#Subtitle { color:#aab8d6; }
        QFrame#Card { background:#0f1a2e; border:1px solid #1f2b46; border-radius:14px; }
        QLineEdit, QComboBox {
            background:#0b1426; border:1px solid #1f2b46; border-radius:10px; padding:8px;
        }
        QComboBox::drop-down { border:0; width:26px; }
        QComboBox::down-arrow {
            image:none;
            border-left:6px solid transparent;
            border-right:6px solid transparent;
            border-top:8px solid #cfe1ff;
            margin-right:8px;
        }
        QPushButton {
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #6d28d9, stop:1 #8b5cf6);
            border:0; border-radius:12px; padding:10px 14px; font-weight:700; color:#ffffff;
        }
        QPushButton:hover { opacity:0.95; }
        QPushButton:disabled { background:#2a3350; color:#9aa7c6; }
        QProgressBar { background:#0b1426; border:1px solid #1f2b46; border-radius:10px; text-align:center; }
        QProgressBar::chunk { background:#8b5cf6; border-radius:10px; }
        QCheckBox { spacing:8px; }
    """,
    "Graphite": """
        QWidget { background:#101010; color:#efefef; font-size:13px; }
        QLabel#Title { font-size:22px; font-weight:800; }
        QLabel#Subtitle { color:#bbbbbb; }
        QFrame#Card { background:#171717; border:1px solid #2a2a2a; border-radius:14px; }
        QLineEdit, QComboBox { background:#121212; border:1px solid #2a2a2a; border-radius:10px; padding:8px; }
        QComboBox::drop-down { border:0; width:26px; }
        QComboBox::down-arrow { image:none; border-left:6px solid transparent; border-right:6px solid transparent; border-top:8px solid #eaeaea; margin-right:8px; }
        QPushButton { background:#2d6cdf; border:0; border-radius:12px; padding:10px 14px; font-weight:700; color:#fff; }
        QProgressBar { background:#121212; border:1px solid #2a2a2a; border-radius:10px; text-align:center; }
        QProgressBar::chunk { background:#2d6cdf; border-radius:10px; }
    """,
    "Light": """
        QWidget { background:#f6f7fb; color:#12131a; font-size:13px; }
        QLabel#Title { font-size:22px; font-weight:800; }
        QLabel#Subtitle { color:#55607a; }
        QFrame#Card { background:#ffffff; border:1px solid #dfe5f1; border-radius:14px; }
        QLineEdit, QComboBox { background:#ffffff; border:1px solid #dfe5f1; border-radius:10px; padding:8px; }
        QComboBox::drop-down { border:0; width:26px; }
        QComboBox::down-arrow { image:none; border-left:6px solid transparent; border-right:6px solid transparent; border-top:8px solid #2b61ff; margin-right:8px; }
        QPushButton { background:#2b61ff; border:0; border-radius:12px; padding:10px 14px; font-weight:700; color:#fff; }
        QProgressBar { background:#ffffff; border:1px solid #dfe5f1; border-radius:10px; text-align:center; }
        QProgressBar::chunk { background:#2b61ff; border-radius:10px; }
    """,
}


# ---------- text pools ----------
SLOGANS = [
    "Красивые", "Крутые", "Стильные", "Модные", "Молодёжные",
    "Дизайнерские", "Эффектные", "Трендовые", "Лаконичные",
    "Яркие", "Современные", "Премиальные", "Универсальные",
    "Актуальные", "Выразительные", "Элегантные", "Минималистичные",
    "Смелые", "Классные", "Городские", "Лёгкие", "Комфортные",
    "Популярные", "Эксклюзивные", "Фирменные", "Невероятные",
    "Супер-стильные", "Изящные", "Брутальные", "Ультрамодные",
]

SUN_TERMS = ["солнцезащитные очки", "солнечные очки"]

SCENARIOS = [
    "город", "путешествия", "отпуск", "прогулки", "вождение",
    "пляж", "активный отдых", "повседневные дела", "поездки", "летние мероприятия"
]

HOOKS_PREMIUM = [
    "Это тот самый аксессуар, который собирает образ в одну линию — спокойно, дорого и уверенно.",
    "Лёгкий акцент, который выглядит «брендово» без лишнего шума.",
    "Форма и посадка продуманы так, чтобы очки смотрелись гармонично в любой стилизации.",
]
HOOKS_NEUTRAL = [
    "Модель легко вписывается в повседневный гардероб и подчёркивает индивидуальность.",
    "Очки комфортны на каждый день и подходят под разные стили.",
    "Надёжный вариант на сезон: удобно, практично и стильно.",
]
HOOKS_MASS = [
    "Удобные очки на каждый день: сочетаются с одеждой и не перегружают образ.",
    "Хороший выбор, если хочется стильный аксессуар без сложностей.",
    "Берёшь — и сразу готовый лук: просто и со вкусом.",
]
HOOKS_SOCIAL = [
    "Тот самый вайб для фото и сторис — выглядит дорого и актуально 😎",
    "С этими очками образ становится «как в подборках» — просто надень и готово.",
    "Лёгкий апгрейд стиля: лайк за форму, лайк за настроение ✨",
]

ENDS = [
    "Отличный вариант на тёплый сезон: стильно, удобно и практично.",
    "Подходят и для города, и для отдыха — аккуратный акцент в образе.",
    "Выбирай как обновление гардероба к сезону — очки легко комбинируются с одеждой.",
]

SEO_KEYS = [
    "очки солнцезащитные", "солнечные очки", "солнцезащитные очки",
    "брендовые очки", "модные очки", "очки женские", "очки мужские",
    "инста очки", "очки из tiktok"
]

# WB strict (убираем абсолюты/обещания/стоп-фразы)
STRICT_DROP = [
    "лучшие", "самые лучшие", "идеальные", "100%", "гарантия", "гарантируем",
    "вылечит", "лечит", "абсолютно", "безусловно", "никогда", "всегда",
]

# WB safe (замена риск-слов — можешь дополнять)
SAFE_REPLACE = {
    "реплика": "стиль в духе бренда",
    "копия": "вдохновлённый дизайн",
    "люкс": "премиальный стиль",
}


def _cut_no_break_words(text: str, limit: int) -> str:
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    return text[:limit].rsplit(" ", 1)[0].strip()


def _uniq(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def normalize_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("-", " ").replace("&", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_brands_ru_map(data_dir: str) -> dict:
    p = Path(data_dir) / "brands_ru.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def brand_ru(brand_lat: str, brand_map: dict) -> str:
    key = normalize_key(brand_lat)
    return (brand_map.get(key) or brand_lat).strip()


def choose_hook(style: str) -> str:
    if style == "premium":
        return random.choice(HOOKS_PREMIUM)
    if style == "mass":
        return random.choice(HOOKS_MASS)
    if style == "social":
        return random.choice(HOOKS_SOCIAL)
    return random.choice(HOOKS_NEUTRAL)


def seo_pack(seo_level: str) -> list[str]:
    base = _uniq(SEO_KEYS)
    random.shuffle(base)
    if seo_level == "low":
        return base[:2]
    if seo_level == "high":
        return base[:6]
    return base[:4]


def apply_safe(text: str) -> str:
    t = text
    for a, b in SAFE_REPLACE.items():
        t = re.sub(rf"\b{re.escape(a)}\b", b, t, flags=re.IGNORECASE)
    return t


def apply_strict(text: str) -> str:
    t = text
    for w in STRICT_DROP:
        t = re.sub(rf"\b{re.escape(w)}\b", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t


def generate_title(brand_lat: str, shape: str, lens: str, brand_map: dict) -> str:
    # RU brand only in title
    b_ru = brand_ru(brand_lat, brand_map)

    parts = [random.choice(SLOGANS), random.choice(SUN_TERMS)]

    # бренд рандомно: 50% как ты просил
    if random.random() < 0.5:
        parts.append(b_ru)

    # форма/линзы — тоже рандомно, но без ломания слов
    if shape and random.random() < 0.6:
        parts.append(shape)
    if lens and random.random() < 0.5:
        parts.append(lens)

    title = " ".join([p for p in parts if p]).strip()
    title = re.sub(r"\s{2,}", " ", title)
    title = title[:1].upper() + title[1:]
    return _cut_no_break_words(title, TITLE_MAX)


def gender_phrase(mode: str) -> str:
    if mode == "Женские":
        return "женские"
    if mode == "Мужские":
        return "мужские"
    if mode == "Унисекс":
        return "унисекс"
    # Auto — нейтрально
    return ""


def description_length_target(desc_length: str) -> int:
    if desc_length == "short":
        return 650
    if desc_length == "long":
        return 1500
    return 1000


def generate_description(
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    style: str,
    seo_level: str,
    gender_mode: str,
) -> str:
    g = gender_phrase(gender_mode)
    hook = choose_hook(style)

    scen = random.sample(SCENARIOS, 4)
    keys = seo_pack(seo_level)

    # без меток "Сценарии: / Линзы:" и т.п. — просто живой текст
    pieces = []
    pieces.append(hook)

    if brand_lat:
        pieces.append(f"Очки {brand_lat} — {('' if not g else g + ' ')}аксессуар, который легко носить каждый день.")
    else:
        pieces.append(f"{('' if not g else g.capitalize() + ' ')}очки — удобный аксессуар на сезон.")

    # форма/линзы вплетаем естественно
    if shape:
        pieces.append(f"Форма {shape} смотрится актуально и помогает сбалансировать черты лица.")
    if lens:
        pieces.append(f"Линзы {lens} дают комфорт при ярком солнце и подходят для активного дня.")

    if collection:
        pieces.append(f"Модель хорошо заходит на сезон {collection} — для города и отдыха.")
    pieces.append(f"Идеи, куда носить: {', '.join(scen)}.")
    pieces.append(random.choice(ENDS))

    # SEO ключи — мягко, одной строкой в конце
    pieces.append(" ".join(keys) + ".")

    text = " ".join(pieces)
    text = re.sub(r"\s{2,}", " ", text).strip()
    return _cut_no_break_words(text, DESC_MAX)


def find_header_col(ws, candidates: set[str], header_scan_rows: int = 20):
    # ищем заголовок по тексту в первых строках
    for r in range(1, header_scan_rows + 1):
        for cell in ws[r]:
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            if val in candidates:
                return cell.column, r
    return None, None


def fill_wb_template(
    input_xlsx: str,
    brand_lat: str,
    shape: str,
    lens: str,
    collection: str,
    style: str = "premium",
    desc_length: str = "medium",
    seo_level: str = "normal",
    gender_mode: str = "Auto",
    wb_safe_mode: bool = True,
    wb_strict: bool = True,
    data_dir: str = "",
    progress_callback=None,
):
    if not input_xlsx:
        raise RuntimeError("Файл XLSX не выбран")

    wb = load_workbook(input_xlsx)
    ws = wb.active

    # ищем колонки
    col_title, header_row = find_header_col(ws, {"наименование", "название"})
    col_desc, header_row2 = find_header_col(ws, {"описание", "description"})

    if not col_title or not col_desc:
        raise RuntimeError("Не найдены колонки Наименование и/или Описание")

    header_row = header_row or header_row2 or 1

    # не трогаем первые 4 строки вообще
    start_row = max(header_row + 1, 5)

    # brand map for RU title
    brand_map = load_brands_ru_map(data_dir) if data_dir else {}

    # защита от дублей в рамках одного прогона
    used_titles = set()
    used_desc = set()

    total_rows = ws.max_row - start_row + 1
    if total_rows <= 0:
        raise RuntimeError("Нет строк для заполнения (после заголовка)")

    processed = 0

    for r in range(start_row, ws.max_row + 1):
        # титул: пробуем несколько раз, чтобы не повторялся
        for _ in range(20):
            t = generate_title(brand_lat, shape, lens, brand_map)
            if t not in used_titles:
                used_titles.add(t)
                break

        d = generate_description(brand_lat, shape, lens, collection, style, seo_level, gender_mode)

        # описания тоже стараемся не повторять
        tries = 0
        while d in used_desc and tries < 10:
            d = generate_description(brand_lat, shape, lens, collection, style, seo_level, gender_mode)
            tries += 1
        used_desc.add(d)

        if wb_safe_mode:
            t = apply_safe(t)
            d = apply_safe(d)
        if wb_strict:
            t = apply_strict(t)
            d = apply_strict(d)

        ws.cell(row=r, column=col_title).value = t
        ws.cell(row=r, column=col_desc).value = d

        processed += 1
        if progress_callback:
            progress_callback((processed / total_rows) * 100)

    out_path = str(Path(input_xlsx).with_name(Path(input_xlsx).stem + "_ready.xlsx"))
    wb.save(out_path)
    return out_path, processed
